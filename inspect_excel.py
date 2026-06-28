import openpyxl

wb = openpyxl.load_workbook(
    r'c:\Users\AAvenido\Documents\Budget Tracker App\BudgetTracker_Final_Update.xlsm',
    keep_vba=True
)

print("=== SHEETS ===")
print(wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== SHEET: {sheet_name} ===")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max Row: {ws.max_row}, Max Col: {ws.max_column}")
    print("\n--- First 60 rows ---")
    for row in ws.iter_rows(min_row=1, max_row=60, values_only=True):
        if any(cell is not None for cell in row):
            print(row)
