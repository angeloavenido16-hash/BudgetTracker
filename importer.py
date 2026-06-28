"""
importer.py  –  One-time import from the original Excel file into SQLite.
Run standalone:  python importer.py
"""

import os
import openpyxl
from datetime import datetime
import database as db

EXCEL_PATH = os.path.join(os.path.dirname(__file__),
                          "BudgetTracker_Final_Update.xlsm")

# ── Sheets that represent a distinct "salary cutoff" period ───────────────
SALARY_PATTERN = [  # sheet names that match MM-DD-YY
    "02-27-25","03-13-25","04-10-25","04-24-25","05-08-25","05-22-25",
    "06-05-25","06-19-25","07-03-25","07-17-25","07-31-25","08-14-25",
    "08-28-25","09-11-25","09-25-25","10-09-25","10-23-25","11-06-25",
    "11-20-25","12-04-25","12-18-25","01-01-26","01-15-26","01-29-26",
    "02-12-26","02-26-26","03-12-26","03-26-26","04-09-26","04-23-26",
    "05-07-26","05-21-26","06-04-26","06-18-26",
]

# ── Special / non-salary funds ────────────────────────────────────────────
# Format: sheet_name -> (display_name, fund_type, total_amount_key)
# total_amount_key must match EXACTLY the string in Column A of "Total Amount" sheet
SPECIAL_FUNDS = {
    "Maya":              ("Maya Account",              "other", "Maya"),
    "Maya Expenses 2":   ("Maya Account 2",            "other", "Maya Expenses 2"),
    "ESPP1":             ("ESPP (Bora Trip)",           "espp",  "ESPP1"),
    "ESPP Expenses 2":   ("ESPP Expenses 2",            "espp",  "ESPP Expenses 2"),
    "13.5 Month Pay":    ("13.5 Month Pay (FY25)",      "bonus", "13.5 Month Pay"),
    "13thFY25":          ("13th Month Pay (FY25)",      "bonus", "13thFY25"),
    "13.5 Month 2026":   ("13.5 Month Pay (FY26)",      "bonus", "13.5 Month 2026"),
    "CorpBonus-6-05-25": ("Corporate Bonus 1Q/2Q25",   "bonus", "Corporate Bonus 1Q25/2Q25"),
    "Corp34Q25":         ("Corporate Bonus 3Q/4Q25",   "bonus", "Corp34Q25"),
    "Corp12Q26":         ("Corporate Bonus 1Q/2Q26",   "bonus", "Corp12Q26"),
}

# ── Categories (from Dropdown Expenses sheet + common ones) ───────────────
KNOWN_CATEGORIES = [
    "24 Chicken","Accommodation","Alta Vista (Breakfast)","Badminton",
    "Baggage","Bar","Batuta Bar","Bayad","BDO Credit Card","Birthday",
    "BPI Credit Card","Budget","Camiguin","Carry Over","Cashout","Cats",
    "CDC Cafe","Chocolate","Chowking","Clothes","Coco Mama",
    "Correction Factor","Date with Babi","Dinner Date","Dinner Out",
    "Disney","Dooki","Dunkin","Electric Bill","Emergency","Figaro","Food",
    "Games","Gas","Gcash","General Cleaning","Gift","Groceries","Haircut",
    "HOHO Bora","House","Interest","Island Chicken Inasal",
    "Jasper's Tapsilog","Jollibee","KFC","Kitchen City","Klook",
    "Laguna Budget","Le Andres Café","Load","LoR","Lunch Out",
    "Mandarin Spa","Massage","Mcdo","Netflix","OM Bar","Order",
    "Outing","Pag-ibig","Payment","Personal Care","Phone",
    "Pickuop Coffee","Plane Ticket","Popeye's","PS4","PSP","Ramen Nagi",
    "Received","Refund","Rent","Saboria","Savings","SM Groceries","Spa",
    "Spice Birds","Spotify","Staycation","Switch 2","Tax",
    "Ticket to BXU","Ticket to MNL","Ticket to MPH","Tiktok Shop",
    "Transfer","Transfer Fee","Two Seasons","Utang","Vape",
    "Water Bill","Watsons","Wifi","Withdraw","Withdraw Fee","Bayad",
]


def _parse_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
            try:
                return datetime.strptime(val.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
    return None


def _parse_amount(val) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        # simple arithmetic like =1000+500
        try:
            cleaned = val.lstrip("=").replace(",", "")
            return float(eval(cleaned))   # safe: only +/-/* on numbers
        except Exception:
            pass
    return None


def _sheet_to_cutoff_label(sheet_name: str) -> str:
    """Convert '02-27-25' -> '02/27/2025 Salary'."""
    parts = sheet_name.split("-")
    if len(parts) == 3:
        mm, dd, yy = parts
        year = "20" + yy
        return f"{mm}/{dd}/{year} Salary"
    return sheet_name + " Salary"


def _sheet_to_date(sheet_name: str) -> str | None:
    parts = sheet_name.split("-")
    if len(parts) == 3:
        mm, dd, yy = parts
        return f"20{yy}-{mm}-{dd}"
    return None


def import_from_excel(status_callback=None):
    """
    Import all data from the Excel file into the SQLite DB.
    Clears existing data first.
    status_callback(msg) is called with progress messages if provided.
    """

    def log(msg):
        if status_callback:
            status_callback(msg)
        else:
            print(msg)

    # ── Wipe existing data ────────────────────────────────────────────────
    with db.get_connection() as conn:
        conn.execute("DELETE FROM transactions")
        conn.execute("DELETE FROM funds")
        conn.execute("DELETE FROM expense_categories")
        conn.execute("DELETE FROM sqlite_sequence"  # reset autoincrement
                     " WHERE name IN ('transactions','funds','expense_categories')")
        conn.commit()

    # ── Seed categories ───────────────────────────────────────────────────
    db.seed_categories(KNOWN_CATEGORIES)
    log("✔ Categories seeded")

    # ── Open workbook ─────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(EXCEL_PATH, keep_vba=True, data_only=True)

    # ── Grab Total Amount row amounts ─────────────────────────────────────
    # Column B = income amount per fund row
    total_ws = wb["Total Amount"]
    fund_amounts = {}   # label -> amount
    for row in total_ws.iter_rows(min_row=1, max_row=50, values_only=True):
        label, amount = row[0], row[1]
        if label and isinstance(amount, (int, float)):
            fund_amounts[str(label).strip()] = float(amount)

    # ── Import salary sheets ──────────────────────────────────────────────
    for sheet_name in SALARY_PATTERN:
        if sheet_name not in wb.sheetnames:
            log(f"  ⚠  Sheet '{sheet_name}' not found, skipping.")
            continue

        label       = _sheet_to_cutoff_label(sheet_name)
        cutoff_date = _sheet_to_date(sheet_name)
        amount      = fund_amounts.get(label, 0.0)

        fund_id = db.add_fund(
            name=label,
            fund_type="salary",
            amount=amount,
            cutoff_date=cutoff_date,
        )

        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        txn_count = 0
        for row in rows:
            if not row or not any(c is not None for c in row):
                continue
            category = row[0]
            amount_v = _parse_amount(row[1])
            date_v   = _parse_date(row[2])
            remarks  = str(row[3]).strip() if row[3] else None

            if not category or amount_v is None:
                continue
            # add category to list if new
            db.add_category(str(category))
            db.add_transaction(
                fund_id=fund_id,
                category=str(category),
                amount=amount_v,
                txn_date=date_v,
                remarks=remarks,
            )
            txn_count += 1

        log(f"  ✔ Salary {sheet_name}  ({txn_count} transactions, ₱{amount:,.2f})")

    # ── Import special/bonus sheets ───────────────────────────────────────
    for sheet_name, (display_name, fund_type, total_key) in SPECIAL_FUNDS.items():
        if sheet_name not in wb.sheetnames:
            log(f"  ⚠  Sheet '{sheet_name}' not found, skipping.")
            continue

        # Look up amount using the exact Total Amount sheet key
        amount = fund_amounts.get(total_key, 0.0)

        # cutoff date from cell F1
        ws       = wb[sheet_name]
        first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        cutoff_date = _parse_date(first_row[6]) if len(first_row) > 6 else None

        fund_id = db.add_fund(
            name=display_name,
            fund_type=fund_type,
            amount=amount,
            cutoff_date=cutoff_date,
        )

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        txn_count = 0
        for row in rows:
            if not row or not any(c is not None for c in row):
                continue
            category = row[0]
            amount_v = _parse_amount(row[1])
            date_v   = _parse_date(row[2])
            remarks  = str(row[3]).strip() if row[3] else None

            if not category or amount_v is None:
                continue
            db.add_category(str(category))
            db.add_transaction(
                fund_id=fund_id,
                category=str(category),
                amount=amount_v,
                txn_date=date_v,
                remarks=remarks,
            )
            txn_count += 1

        log(f"  ✔ Special '{display_name}'  ({txn_count} transactions, ₱{amount:,.2f})")

    # ── BPI balance (from Total Amount M1) ───────────────────────────────
    try:
        bpi_cell = total_ws["M1"].value
        if isinstance(bpi_cell, (int, float)):
            db.update_bpi_balance(float(bpi_cell))
            log(f"  ✔ BPI balance set to ₱{bpi_cell:,.2f}")
    except Exception:
        pass

    log("\n✅ Import complete!")


if __name__ == "__main__":
    db.initialize_db()
    import_from_excel()
